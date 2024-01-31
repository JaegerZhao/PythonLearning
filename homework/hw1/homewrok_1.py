'''
将2012_2021.csv文件中的数据，整理到excel中； 要求：

根据年份创建对应的sheet，
将年份对应的数据保存到对应的sheet中，
使用面向对象实现
文件打开编码格式：utf-8-sig
'''
import csv
from openpyxl import Workbook

class CSVToExcel:
    def __init__(self, csv_file, encoding='utf-8-sig'):
        self.csv_file = csv_file
        self.encoding = encoding
        
    def convert(self):
        workbook = Workbook()
        with open(self.csv_file, encoding=self.encoding) as f:
            reader = csv.reader(f)
            headers = next(reader)
            
            for row in reader:
                year = row[1].split('年')[0]
                if year not in workbook.sheetnames:
                    worksheet = workbook.create_sheet(year)
                    worksheet.append(headers)
                else:
                    worksheet = workbook[year]
                worksheet.append(row)
            del workbook['Sheet']
            
        return workbook
                
csv_file = r'data\homework_1.csv'
converter = CSVToExcel(csv_file)
workbook = converter.convert()
workbook.save('homework_1.xlsx')